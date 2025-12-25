"""
===================================================================
Python Excel Automation Script (Single Script – End to End)
===================================================================

WHAT THIS SCRIPT DOES:
---------------------
1. Reads Excel data
2. Removes rows with missing values
3. Calculates total sales amount
4. Adds business status (High / Medium / Low)
5. Creates a summary sheet
6. Applies conditional formatting (colors)
7. Creates Excel charts (Bar + Pie)
8. Saves final professional Excel file

INPUT  FILE : sales_data.xlsx
OUTPUT FILE : Final_Excel_Report.xlsx
===================================================================
"""

# ============================================================
# IMPORT REQUIRED LIBRARIES
# ============================================================

# pandas:
# Used to read Excel files and perform data operations
import pandas as pd

# load_workbook:
# Used to open an Excel file again for formatting & charts
from openpyxl import load_workbook

# PatternFill:
# Used to apply background colors to Excel cells
from openpyxl.styles import PatternFill

# Chart classes:
# Used to create Excel charts using Python
from openpyxl.chart import BarChart, PieChart, Reference


# ============================================================
# FILE CONFIGURATION
# ============================================================

# Name of the input Excel file
INPUT_FILE = "sales_data.xlsx"

# Name of the output Excel file
OUTPUT_FILE = "Final_Excel_Report.xlsx"


# ============================================================
# STEP 1: READ EXCEL FILE
# ============================================================

# Read Excel file and load it into a DataFrame
# DataFrame = Excel table inside Python
df = pd.read_excel(INPUT_FILE)

# Print confirmation message
print("📥 Excel file loaded successfully")

# At this point:
# df contains columns: Date, Product, Quantity, Price


# ============================================================
# STEP 2: DATA CLEANING
# ============================================================

# dropna() removes rows with missing (empty) values
# inplace=True means original DataFrame is modified
df.dropna(inplace=True)

# Print confirmation message
print("🧹 Missing rows removed")

# Result:
# Rows where Quantity or Price was empty are deleted


# ============================================================
# STEP 3: ADD CALCULATED COLUMN
# ============================================================

# Create a new column called "Total Amount"
# Formula: Quantity * Price
df["Total Amount"] = df["Quantity"] * df["Price"]

# Print confirmation message
print("🧮 Total Amount column created")

# Example:
# Quantity = 2, Price = 50000 → Total Amount = 100000


# ============================================================
# STEP 4: ADD STATUS COLUMN (BUSINESS LOGIC)
# ============================================================

def sales_status(amount):
    """
    This function decides the sales category
    based on Total Amount value.

    BUSINESS RULES:
    - amount >= 100000 → High
    - amount >= 50000  → Medium
    - amount < 50000   → Low
    """

    # If sales amount is very high
    if amount >= 100000:
        return "High"

    # If sales amount is medium
    elif amount >= 50000:
        return "Medium"

    # If sales amount is low
    else:
        return "Low"


# apply() runs the function on each row value
# It applies sales_status() to every Total Amount
df["Sales Status"] = df["Total Amount"].apply(sales_status)

# Print confirmation message
print("🏷 Sales Status column added")


# ============================================================
# STEP 5: CREATE SUMMARY REPORT
# ============================================================

# Create a new DataFrame for summary information
summary_df = pd.DataFrame({

    # First column: Description
    "Metric": [
        "Total Revenue",
        "Total Orders",
        "High Sales Orders",
        "Medium Sales Orders",
        "Low Sales Orders"
    ],

    # Second column: Calculated values
    "Value": [
        df["Total Amount"].sum(),                    # Sum of all sales
        len(df),                                     # Total number of records
        (df["Sales Status"] == "High").sum(),        # Count of High sales
        (df["Sales Status"] == "Medium").sum(),      # Count of Medium sales
        (df["Sales Status"] == "Low").sum()          # Count of Low sales
    ]
})

# Print confirmation message
print("📊 Summary report generated")


# ============================================================
# STEP 6: WRITE DATA TO EXCEL
# ============================================================

# ExcelWriter allows writing multiple sheets into one Excel file
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

    # Write cleaned sales data to first sheet
    df.to_excel(
        writer,
        sheet_name="Sales Data",     # Sheet name
        index=False                  # Do not write index column
    )

    # Write summary data to second sheet
    summary_df.to_excel(
        writer,
        sheet_name="Summary",
        index=False
    )

# Print confirmation message
print("💾 Data written to Excel")


# ============================================================
# STEP 7: APPLY CONDITIONAL FORMATTING
# ============================================================

# Load the newly created Excel file
wb = load_workbook(OUTPUT_FILE)

# Select the "Sales Data" sheet
ws = wb["Sales Data"]

# ------------------------------------------------------------
# DEFINE CELL COLORS USING HEX CODES
# ------------------------------------------------------------

# Light green color for High sales
high_fill = PatternFill(
    start_color="C6EFCE",   # HEX code for green
    end_color="C6EFCE",     # Same color for full fill
    fill_type="solid"       # Solid background
)

# Light yellow color for Medium sales
medium_fill = PatternFill(
    start_color="FFEB9C",   # HEX code for yellow
    end_color="FFEB9C",
    fill_type="solid"
)

# Light red color for Low sales
low_fill = PatternFill(
    start_color="FFC7CE",   # HEX code for red
    end_color="FFC7CE",
    fill_type="solid"
)

# ------------------------------------------------------------
# APPLY COLORS ROW BY ROW
# ------------------------------------------------------------

# Start loop from row 2 (row 1 contains headers)
for row in range(2, ws.max_row + 1):

    # Access Sales Status cell (Column F)
    status_cell = ws[f"F{row}"]

    # Apply background color based on status value
    if status_cell.value == "High":
        status_cell.fill = high_fill

    elif status_cell.value == "Medium":
        status_cell.fill = medium_fill

    elif status_cell.value == "Low":
        status_cell.fill = low_fill

# Print confirmation message
print("🎨 Conditional formatting applied")


# ============================================================
# STEP 8: CREATE EXCEL CHARTS
# ============================================================

# Create a new sheet named "Charts"
chart_sheet = wb.create_sheet(title="Charts")


# ------------------------------------------------------------
# BAR CHART: PRODUCT vs TOTAL AMOUNT
# ------------------------------------------------------------

# Create bar chart object
bar_chart = BarChart()

# Set chart title
bar_chart.title = "Product Wise Total Sales"

# Set X and Y axis titles
bar_chart.x_axis.title = "Product"
bar_chart.y_axis.title = "Total Amount"

# Select Total Amount column (Column E)
data = Reference(
    ws,
    min_col=5,          # Column E
    min_row=1,          # Include header
    max_row=ws.max_row
)

# Select Product column (Column B)
categories = Reference(
    ws,
    min_col=2,          # Column B
    min_row=2,          # Skip header
    max_row=ws.max_row
)

# Add data and categories to chart
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)

# Add bar chart to Charts sheet
chart_sheet.add_chart(bar_chart, "A1")


# ------------------------------------------------------------
# PIE CHART: SALES STATUS DISTRIBUTION
# ------------------------------------------------------------

# Create pie chart object
pie_chart = PieChart()

# Set chart title
pie_chart.title = "Sales Status Distribution"

# Labels (High / Medium / Low)
labels = Reference(
    wb["Summary"],
    min_col=1,
    min_row=4,
    max_row=6
)

# Values for pie chart
data = Reference(
    wb["Summary"],
    min_col=2,
    min_row=4,
    max_row=6
)

# Add data and labels
pie_chart.add_data(data, titles_from_data=False)
pie_chart.set_categories(labels)

# Add pie chart to Charts sheet
chart_sheet.add_chart(pie_chart, "A20")


# ============================================================
# FINAL SAVE
# ============================================================

# Save the Excel file with all changes
wb.save(OUTPUT_FILE)

# Final confirmation messages
print("📊 Charts created successfully")
print("✅ Excel Automation Completed Successfully")
